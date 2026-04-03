#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
can_mqtt_excel.py  —  CAN-MQTT 规则 Excel 工具

功能：
  1. 生成示例 Excel 模板（含说明行）
     python3 can_mqtt_excel.py gen [output.xlsx]

  2. 将 Excel 文件转换为设备 JSON 规则
     python3 can_mqtt_excel.py to_json input.xlsx [output.json]

  3. 将设备 JSON 规则导出为 Excel
     python3 can_mqtt_excel.py from_json input.json [output.xlsx]

Excel 列定义（与 can_mqtt_rule_t 对应）：
  A  规则ID         (rule_id)       必填，唯一字符串，如 "bms_soc_001"
  B  规则名称       (name)          选填，便于阅读
  C  启用           (enabled)       必填，TRUE / FALSE
  D  优先级         (priority)      必填，整数，数字越大优先级越高
  E  CAN通道        (channel)       必填，"can0" / "can1" / "any"
  F  CAN ID (十六进制) (can_id)     必填，如 "0x1A5" 或 "1A5"；填 "ANY" 则匹配任意帧
  G  扩展帧         (is_extended)   必填，TRUE=29位ID / FALSE=11位ID
  H  报文名称       (message_name)  选填，用于 MQTT topic 模板替换
  I  信号名称       (signal_name)   必填，用于 MQTT topic 模板替换
  J  起始位         (start_bit)     必填，Intel=LSB位号 / Motorola=MSB位号，0起
  K  位长度         (bit_length)    必填，1-64
  L  字节序         (byte_order)    必填，"Intel" 或 "Motorola"
  M  有符号         (is_signed)     必填，TRUE / FALSE
  N  系数           (factor)        必填，物理值 = 原始值 * factor + offset
  O  偏移           (offset)        必填，物理值 = 原始值 * factor + offset
  P  单位           (unit)          选填，如 "V" "A" "%" "°C"
  Q  MQTT主题模板   (topic_template) 必填，支持占位符：
                                     {topic_prefix}  设备配置的前缀
                                     {device_id}     设备 ID
                                     {message_name}  H列
                                     {signal_name}   I列
                                     {channel}       CAN通道
  R  载荷模式       (payload_mode)  必填，"json" 或 "raw"
  S  QoS            (qos)           必填，0 / 1 / 2
  T  保留消息       (retain)        必填，TRUE / FALSE
"""

import sys
import os
import json
import uuid
import time
import struct

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, Protection
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
except ImportError:
    sys.exit("请先安装 openpyxl：pip3 install openpyxl")

# ────────────────────────────────────────────────────────────────────────────
#  颜色常量
# ────────────────────────────────────────────────────────────────────────────
C_HEADER   = "1E3A5F"   # 深蓝色
C_GROUP    = "2E4D7B"   # 中蓝色
C_EXAMPLE  = "0F3460"   # 示例行底色
C_LIGHT    = "D9E8FF"   # 浅蓝用于验证下拉
C_WHITE    = "FFFFFF"
C_YELLOW   = "FFF3CD"   # 说明行
C_RED      = "FF4444"   # 必填
C_GREEN    = "27AE60"   # 可选

# ────────────────────────────────────────────────────────────────────────────
#  列定义
# ────────────────────────────────────────────────────────────────────────────
COLUMNS = [
    # (列标题,            内部字段,         宽度, 必填, 说明)
    ("规则ID",           "rule_id",         18, True,  "唯一字符串，如 bms_soc_001"),
    ("规则名称",          "name",            22, False, "便于阅读的描述性名称"),
    ("启用",             "enabled",         8,  True,  "TRUE=启用 FALSE=禁用"),
    ("优先级",           "priority",        8,  True,  "整数，越大越先匹配，默认 0"),
    ("CAN通道",          "channel",         10, True,  "can0 / can1 / any"),
    ("CAN ID(十六进制)", "can_id",           14, True,  "如 1A5 或 0x1A5；填 ANY 匹配所有"),
    ("扩展帧(29位)",     "is_extended",     12, True,  "TRUE=29位扩展帧 FALSE=11位标准帧"),
    ("报文名称",          "message_name",    18, False, "DBC 报文名，用于 topic 模板"),
    ("信号名称",          "signal_name",     18, True,  "DBC 信号名，用于 topic 模板"),
    ("起始位",           "start_bit",       9,  True,  "Intel=LSB位号; Motorola=MSB位号 (0起)"),
    ("位长度",           "bit_length",      8,  True,  "1~64"),
    ("字节序",           "byte_order",      10, True,  "Intel 或 Motorola"),
    ("有符号",           "is_signed",       8,  True,  "TRUE=有符号整数 FALSE=无符号"),
    ("系数 factor",      "factor",          12, True,  "物理值 = 原始值 × factor + offset"),
    ("偏移 offset",      "offset",          12, True,  "物理值 = 原始值 × factor + offset"),
    ("单位",             "unit",            8,  False, "如 V A % °C rpm"),
    ("MQTT主题模板",     "topic_template",  40, True,
     "支持占位符: {topic_prefix} {device_id} {message_name} {signal_name} {channel}"),
    ("载荷模式",          "payload_mode",    10, True,  "json = 完整JSON; raw = 纯数值字符串"),
    ("QoS",             "qos",             6,  True,  "0/1/2"),
    ("保留消息",          "retain",          8,  True,  "TRUE=Retained FALSE=普通"),
]

# ────────────────────────────────────────────────────────────────────────────
#  示例数据行（BMS 典型信号）
# ────────────────────────────────────────────────────────────────────────────
EXAMPLES = [
    {
        "rule_id": "bms_total_voltage",
        "name": "BMS 总电压",
        "enabled": True,
        "priority": 10,
        "channel": "can0",
        "can_id": "0x1A5",
        "is_extended": False,
        "message_name": "BMS_Status",
        "signal_name": "TotalVoltage",
        "start_bit": 0,
        "bit_length": 16,
        "byte_order": "Intel",
        "is_signed": False,
        "factor": 0.1,
        "offset": 0.0,
        "unit": "V",
        "topic_template": "{topic_prefix}/{device_id}/bms/voltage",
        "payload_mode": "json",
        "qos": 1,
        "retain": False,
    },
    {
        "rule_id": "bms_current",
        "name": "BMS 电流",
        "enabled": True,
        "priority": 10,
        "channel": "can0",
        "can_id": "0x1A5",
        "is_extended": False,
        "message_name": "BMS_Status",
        "signal_name": "Current",
        "start_bit": 16,
        "bit_length": 16,
        "byte_order": "Intel",
        "is_signed": True,
        "factor": 0.1,
        "offset": -3000.0,
        "unit": "A",
        "topic_template": "{topic_prefix}/{device_id}/bms/current",
        "payload_mode": "json",
        "qos": 1,
        "retain": False,
    },
    {
        "rule_id": "bms_soc",
        "name": "BMS SOC",
        "enabled": True,
        "priority": 10,
        "channel": "can0",
        "can_id": "0x1A6",
        "is_extended": False,
        "message_name": "BMS_SOC",
        "signal_name": "SOC",
        "start_bit": 0,
        "bit_length": 8,
        "byte_order": "Intel",
        "is_signed": False,
        "factor": 1.0,
        "offset": 0.0,
        "unit": "%",
        "topic_template": "{topic_prefix}/{device_id}/bms/soc",
        "payload_mode": "json",
        "qos": 1,
        "retain": True,
    },
    {
        "rule_id": "bms_cell_temp_max",
        "name": "BMS 最高单体温度",
        "enabled": True,
        "priority": 5,
        "channel": "can1",
        "can_id": "0x1B0",
        "is_extended": False,
        "message_name": "BMS_Temperature",
        "signal_name": "CellTempMax",
        "start_bit": 0,
        "bit_length": 8,
        "byte_order": "Motorola",
        "is_signed": True,
        "factor": 1.0,
        "offset": -40.0,
        "unit": "°C",
        "topic_template": "{topic_prefix}/{device_id}/bms/temp/cell_max",
        "payload_mode": "json",
        "qos": 0,
        "retain": False,
    },
    {
        "rule_id": "motor_rpm",
        "name": "电机转速",
        "enabled": True,
        "priority": 8,
        "channel": "any",
        "can_id": "0x210",
        "is_extended": False,
        "message_name": "Motor_Status",
        "signal_name": "MotorRPM",
        "start_bit": 8,
        "bit_length": 16,
        "byte_order": "Intel",
        "is_signed": True,
        "factor": 1.0,
        "offset": 0.0,
        "unit": "rpm",
        "topic_template": "{topic_prefix}/{device_id}/motor/rpm",
        "payload_mode": "raw",
        "qos": 0,
        "retain": False,
    },
]

# ────────────────────────────────────────────────────────────────────────────
#  生成模板
# ────────────────────────────────────────────────────────────────────────────

def _make_thin_border():
    s = Side(style='thin', color='2A3550')
    return Border(left=s, right=s, top=s, bottom=s)


def gen_template(out_path: str):
    wb = openpyxl.Workbook()

    # ── 规则表 ─────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "CAN_MQTT规则"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    header_fill   = PatternFill("solid", fgColor=C_HEADER)
    example_fill  = PatternFill("solid", fgColor=C_EXAMPLE)
    notice_fill   = PatternFill("solid", fgColor=C_YELLOW)
    header_font   = Font(bold=True, color=C_WHITE, size=10)
    notice_font   = Font(color="664D03", size=9, italic=True)
    thin_border   = _make_thin_border()

    # 第1行：说明
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:T1")
    cell = ws["A1"]
    cell.value = (
        "📋  CAN → MQTT 规则导入模板  |  "
        "绿色字段=可选，红色标题=必填  |  "
        "从第3行开始填写规则（示例行可删除）  |  "
        "导入地址: POST http://<服务端IP>:18080/api/rules/import_excel"
    )
    cell.font   = notice_font
    cell.fill   = notice_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # 第2行：列标题
    ws.row_dimensions[2].height = 28
    for col_idx, (title, field, width, required, tip) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=2, column=col_idx, value=title)
        c.font      = Font(bold=True, color=C_WHITE, size=10)
        c.fill      = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border
        if not required:
            c.font = Font(bold=True, color=C_GREEN, size=10)
        # 列宽
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        # 批注
        if tip:
            c.comment = Comment(tip, "模板", width=200, height=60)

    # 数据验证
    _add_validations(ws)

    # 示例数据
    for row_idx, ex in enumerate(EXAMPLES, start=3):
        ws.row_dimensions[row_idx].height = 20
        row_vals = _rule_to_row(ex)
        for col_idx, val in enumerate(row_vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill   = example_fill
            c.font   = Font(color=C_WHITE, size=10)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center")

    # ── 说明表 ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("字段说明")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 50

    _write_docs(ws2)

    wb.save(out_path)
    print(f"✅  Excel 模板已生成: {out_path}")
    print(f"   包含 {len(EXAMPLES)} 条示例规则")


def _add_validations(ws):
    TRUE_FALSE_VALS = '"TRUE,FALSE"'
    BYTE_ORDER_VALS = '"Intel,Motorola"'
    CHANNEL_VALS    = '"can0,can1,any"'
    PAYLOAD_VALS    = '"json,raw"'
    QOS_VALS        = '"0,1,2"'

    for key, formula in [
        ("C", TRUE_FALSE_VALS),
        ("G", TRUE_FALSE_VALS),
        ("M", TRUE_FALSE_VALS),
        ("T", TRUE_FALSE_VALS),
        ("E", CHANNEL_VALS),
        ("L", BYTE_ORDER_VALS),
        ("R", PAYLOAD_VALS),
        ("S", QOS_VALS),
    ]:
        dv = DataValidation(type="list", formula1=formula, allow_blank=False, showDropDown=False)
        dv.error    = "请从下拉列表中选择"
        dv.errorTitle = "输入无效"
        ws.add_data_validation(dv)
        dv.add(f"{key}3:{key}1000")


def _rule_to_row(r: dict) -> list:
    """dict → 按列顺序的列表值"""
    return [
        r.get("rule_id", ""),
        r.get("name", ""),
        "TRUE" if r.get("enabled", True) else "FALSE",
        r.get("priority", 0),
        r.get("channel", "any"),
        r.get("can_id", ""),
        "TRUE" if r.get("is_extended", False) else "FALSE",
        r.get("message_name", ""),
        r.get("signal_name", ""),
        r.get("start_bit", 0),
        r.get("bit_length", 8),
        r.get("byte_order", "Intel"),
        "TRUE" if r.get("is_signed", False) else "FALSE",
        r.get("factor", 1.0),
        r.get("offset", 0.0),
        r.get("unit", ""),
        r.get("topic_template", ""),
        r.get("payload_mode", "json"),
        r.get("qos", 1),
        "TRUE" if r.get("retain", False) else "FALSE",
    ]


def _write_docs(ws):
    ws.title = "字段说明"
    ws["A1"] = "字段名"
    ws["B1"] = "类型"
    ws["C1"] = "说明"
    for c in ["A1","B1","C1"]:
        ws[c].font = Font(bold=True, color="FFFFFF")
        ws[c].fill = PatternFill("solid", fgColor="1E3A5F")
        ws[c].alignment = Alignment(horizontal="center")

    docs = [
        ("rule_id",       "字符串",   "唯一规则ID，建议格式：消息名_信号名，不能重复"),
        ("name",          "字符串",   "人类可读名称，可为中文"),
        ("enabled",       "TRUE/FALSE","FALSE=该规则暂停生效"),
        ("priority",      "整数",     "匹配优先级，数值越大越先执行，默认0"),
        ("channel",       "枚举",     "can0=仅 CAN0  can1=仅 CAN1  any=两个通道都匹配"),
        ("can_id",        "十六进制", "填 0x... 或直接十六进制数字，如 0x1A5 或 1A5；填 ANY 表示匹配所有帧"),
        ("is_extended",   "TRUE/FALSE","TRUE=29位扩展帧（含EFF/RTR/ERR标志）  FALSE=11位标准帧"),
        ("message_name",  "字符串",   "DBC 报文名称，仅用于 MQTT topic 模板占位符替换，可留空"),
        ("signal_name",   "字符串",   "DBC 信号名称，用于 topic 模板占位符 {signal_name}"),
        ("start_bit",     "整数",     (
            "与 DBC/XLS 原始 bit 编号一致（无需手动转换）\n"
            "Intel:    LSB 线性位置 (byte×8 + bit, 0=LSB)\n"
            "Motorola: MSB 的 Motorola bit 编号 (bit0=字节0 MSB, bit7=字节0 LSB)\n"
            "示例: 生命信号在字节7全部8位 → Motorola start_bit=56"
        )),
        ("bit_length",    "整数",     "信号位宽，1~64"),
        ("byte_order",    "枚举",     "Intel = 小端 (LSB first)；Motorola = 大端 (MSB first)"),
        ("is_signed",     "TRUE/FALSE","TRUE=有符号整数（补码）  FALSE=无符号"),
        ("factor",        "浮点",     "物理值 = 原始整数 × factor + offset"),
        ("offset",        "浮点",     "物理值 = 原始整数 × factor + offset"),
        ("unit",          "字符串",   "物理单位，如 V A % °C rpm，可留空"),
        ("topic_template","字符串",   (
            "MQTT 主题模板，可用占位符：\n"
            "  {topic_prefix}  → 设备配置的 MQTT 前缀\n"
            "  {device_id}     → 设备唯一ID\n"
            "  {message_name}  → message_name 列\n"
            "  {signal_name}   → signal_name 列\n"
            "  {channel}       → 实际通道 can0/can1\n"
            "示例: {topic_prefix}/{device_id}/bms/{signal_name}"
        )),
        ("payload_mode",  "枚举",     (
            "json → MQTT 消息体为完整 JSON:\n"
            '  {"signal":"SOC","value":85.0,"unit":"%","ts":1711000000}\n'
            "raw  → MQTT 消息体仅为数值字符串: 85.0"
        )),
        ("qos",           "整数",     "MQTT QoS 等级：0=最多一次  1=至少一次  2=精确一次"),
        ("retain",        "TRUE/FALSE","TRUE=Broker 保留最后一条消息，供订阅者上线后立即收到"),
    ]

    for row_idx, (field, typ, desc) in enumerate(docs, start=2):
        ws.cell(row=row_idx, column=1, value=field).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=typ)
        c = ws.cell(row=row_idx, column=3, value=desc)
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row_idx].height = max(14, desc.count('\n') * 14 + 14)


# ────────────────────────────────────────────────────────────────────────────
#  Excel → JSON 转换
# ────────────────────────────────────────────────────────────────────────────

def _bool(val) -> bool:
    if isinstance(val, bool): return val
    return str(val).strip().upper() in ("TRUE", "1", "YES", "是", "√")


def _int_safe(val, default=0) -> int:
    try: return int(val)
    except: return default


def _float_safe(val, default=0.0) -> float:
    try: return float(val)
    except: return default


def excel_to_rules(xlsx_path: str) -> dict:
    """
    读取 Excel 文件，返回设备 JSON 规则结构。
    兼容：第2行=列标题，第3行起=数据（第1行为说明行，会自动跳过）。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 找到规则 sheet
    ws = None
    for name in wb.sheetnames:
        if "规则" in name or "rule" in name.lower() or name == wb.sheetnames[0]:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    # 找到标题行（第2行 或 检索含"规则ID"的行）
    header_row = 2
    for r in range(1, min(6, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(row=r, column=c).value or "")
            if "规则ID" in v or "rule_id" in v.lower():
                header_row = r
                break

    # 建立 列标题 → 列索引 映射
    col_map: dict[str, int] = {}
    FIELD_ALIASES = {
        "rule_id":       ["规则ID","rule_id","id"],
        "name":          ["规则名称","名称","name"],
        "enabled":       ["启用","enabled"],
        "priority":      ["优先级","priority"],
        "channel":       ["CAN通道","通道","channel"],
        "can_id":        ["CAN ID","can_id","帧ID"],
        "is_extended":   ["扩展帧","is_extended"],
        "message_name":  ["报文名称","message_name","报文"],
        "signal_name":   ["信号名称","signal_name","信号"],
        "start_bit":     ["起始位","start_bit"],
        "bit_length":    ["位长度","bit_length"],
        "byte_order":    ["字节序","byte_order"],
        "is_signed":     ["有符号","is_signed"],
        "factor":        ["系数","factor"],
        "offset":        ["偏移","offset"],
        "unit":          ["单位","unit"],
        "topic_template":["MQTT主题","topic_template","主题模板"],
        "payload_mode":  ["载荷模式","payload_mode"],
        "qos":           ["QoS","qos"],
        "retain":        ["保留消息","retain"],
    }
    for c in range(1, ws.max_column + 1):
        header_val = str(ws.cell(row=header_row, column=c).value or "").strip()
        for field, aliases in FIELD_ALIASES.items():
            for alias in aliases:
                if alias in header_val or header_val in aliases:
                    col_map[field] = c
                    break

    def get_val(row, field):
        if field not in col_map:
            return None
        return ws.cell(row=row, column=col_map[field]).value

    rules = []
    for r in range(header_row + 1, ws.max_row + 1):
        rule_id_val = get_val(r, "rule_id")
        if rule_id_val is None or str(rule_id_val).strip() == "":
            continue
        rule_id = str(rule_id_val).strip()

        # CAN ID 解析
        can_id_raw = str(get_val(r, "can_id") or "").strip()
        match_any_id = can_id_raw.upper() == "ANY" or can_id_raw == ""
        can_id_int = 0
        if not match_any_id:
            try:
                can_id_int = int(can_id_raw.replace("0x","").replace("0X",""), 16)
            except ValueError:
                print(f"  警告 行{r}: CAN ID '{can_id_raw}' 无法解析，已跳过该行")
                continue

        # 字节序 → C代码期望字符串 "big_endian" / "little_endian"
        byte_order_raw = str(get_val(r, "byte_order") or "Intel").strip().lower()
        byte_order_str = "big_endian" if ("motor" in byte_order_raw or "big" in byte_order_raw) else "little_endian"

        # 载荷模式 → C代码期望字符串 "json" / "raw"
        payload_raw = str(get_val(r, "payload_mode") or "json").strip().lower()
        payload_mode_str = "raw" if payload_raw == "raw" else "json"

        channel   = str(get_val(r, "channel") or "any").strip().lower()
        msg_name  = str(get_val(r, "message_name") or "")
        sig_name  = str(get_val(r, "signal_name")  or "")

        rule = {
            "id":       rule_id,
            "name":     str(get_val(r, "name") or rule_id),
            "enabled":  _bool(get_val(r, "enabled") if get_val(r, "enabled") is not None else True),
            "priority": _int_safe(get_val(r, "priority"), 100),
            # ── 匹配条件（嵌套 match 对象，与C代码对齐）──────────────────
            "match": {
                "channel":      channel,
                "can_id":       can_id_int,
                "is_extended":  _bool(get_val(r, "is_extended")),
                "match_any_id": match_any_id,
            },
            # ── 信号来源（嵌套 source 对象）──────────────────────────────
            "source": {
                "type":         "manual_field",
                "message_name": msg_name,
                "signal_name":  sig_name,
            },
            # ── 解码参数（嵌套 decode 对象，字段名与C代码对齐）──────────
            "decode": {
                "mode":       "manual_field",
                "start_bit":  _int_safe(get_val(r, "start_bit"), 0),
                "bit_length": _int_safe(get_val(r, "bit_length"), 8),
                "byte_order": byte_order_str,
                "signed":     _bool(get_val(r, "is_signed")),
                "factor":     _float_safe(get_val(r, "factor"), 1.0),
                "offset":     _float_safe(get_val(r, "offset"), 0.0),
                "unit":       str(get_val(r, "unit") or ""),
            },
            # ── MQTT 发布参数 ─────────────────────────────────────────────
            "mqtt": {
                "topic_template": str(get_val(r, "topic_template") or
                                      "{topic_prefix}/device/{device_id}/{signal_name}"),
                "payload_mode":   payload_mode_str,
                "qos":            _int_safe(get_val(r, "qos"), 0),
                "retain":         _bool(get_val(r, "retain")),
            },
        }
        rules.append(rule)

    return {
        "version":    1,
        "updated_at": int(time.time()),
        "rules":      rules,
    }


# ────────────────────────────────────────────────────────────────────────────
#  JSON → Excel 导出
# ────────────────────────────────────────────────────────────────────────────

def rules_to_excel(rules_obj: dict, out_path: str):
    """将设备 JSON 规则导出为 Excel（方便用户编辑后重新导入）。
    支持嵌套 JSON 格式：match / source / decode / mqtt 子对象。
    """
    rules = rules_obj.get("rules", [])
    ex_list = []
    for rule in rules:
        # 兼容嵌套结构（当前格式）和老式平铺结构
        match  = rule.get("match",  rule)
        src    = rule.get("source", rule)
        dec    = rule.get("decode", {})
        pub    = rule.get("mqtt",   {})

        can_id_int   = match.get("can_id", rule.get("can_id", 0))
        match_any_id = match.get("match_any_id", rule.get("match_any_id", False))
        can_id_str   = "ANY" if match_any_id else f"0x{can_id_int:X}"

        # byte_order: 新格式是字符串 "big_endian"/"little_endian"，老格式是 0/1
        bo_raw = dec.get("byte_order", 0)
        if isinstance(bo_raw, str):
            is_big = "big" in bo_raw.lower() or "motor" in bo_raw.lower()
        else:
            is_big = (bo_raw == 1)

        # signed 字段名：新格式是 "signed"，老格式是 "is_signed"
        is_signed = dec.get("signed", dec.get("is_signed", False))

        # payload_mode：新格式是字符串 "json"/"raw"，老格式是 0/1
        pm_raw = pub.get("payload_mode", 0)
        payload_mode_str = "raw" if (pm_raw == 1 or pm_raw == "raw") else "json"

        ex_list.append({
            "rule_id":        rule.get("id", ""),
            "name":           rule.get("name", ""),
            "enabled":        rule.get("enabled", True),
            "priority":       rule.get("priority", 0),
            "channel":        match.get("channel", rule.get("channel", "any")),
            "can_id":         can_id_str,
            "is_extended":    match.get("is_extended", rule.get("is_extended", False)),
            "message_name":   src.get("message_name", rule.get("message_name", "")),
            "signal_name":    src.get("signal_name",  rule.get("signal_name",  "")),
            "start_bit":      dec.get("start_bit",  0),
            "bit_length":     dec.get("bit_length", 8),
            "byte_order":     "Motorola" if is_big else "Intel",
            "is_signed":      is_signed,
            "factor":         dec.get("factor",  1.0),
            "offset":         dec.get("offset",  0.0),
            "unit":           dec.get("unit",    ""),
            "topic_template": pub.get("topic_template", ""),
            "payload_mode":   payload_mode_str,
            "qos":            pub.get("qos", 1),
            "retain":         pub.get("retain", False),
        })
    # 复用生成函数（临时替换 EXAMPLES）
    global EXAMPLES
    orig = EXAMPLES
    EXAMPLES = ex_list
    gen_template(out_path)
    EXAMPLES = orig
    print(f"   共导出 {len(rules)} 条规则")


# ────────────────────────────────────────────────────────────────────────────
#  CLI 入口
# ────────────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        return

    cmd = args[0].lower()

    if cmd == "gen":
        out = args[1] if len(args) > 1 else "can_mqtt_rules_template.xlsx"
        gen_template(out)

    elif cmd == "to_json":
        if len(args) < 2:
            print("用法: to_json <input.xlsx> [output.json]")
            return
        xlsx = args[1]
        out  = args[2] if len(args) > 2 else xlsx.replace(".xlsx", ".json")
        obj  = excel_to_rules(xlsx)
        with open(out, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)
        print(f"✅  已转换为 JSON: {out}  ({len(obj['rules'])} 条规则)")

    elif cmd == "from_json":
        if len(args) < 2:
            print("用法: from_json <input.json> [output.xlsx]")
            return
        jf  = args[1]
        out = args[2] if len(args) > 2 else jf.replace(".json", "_export.xlsx")
        with open(jf, "r", encoding="utf-8") as f:
            obj = json.load(f)
        rules_to_excel(obj, out)

    else:
        print(f"未知命令: {cmd}")
        print("可用命令: gen / to_json / from_json")


if __name__ == "__main__":
    main()
